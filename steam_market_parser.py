import argparse
import os
import queue
import random
import re
import threading
import time
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from typing import Dict, List, Optional, Set, Tuple
from urllib.parse import quote

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook


# -----------------------------
# Data models
# -----------------------------

@dataclass(frozen=True)
class Proxy:
    host: str
    port: int
    user: str
    password: str

    def as_requests_proxy(self, scheme: str = "http") -> dict:
        u = quote(self.user, safe="")
        p = quote(self.password, safe="")
        proxy_url = f"{scheme}://{u}:{p}@{self.host}:{self.port}"
        return {"http": proxy_url, "https": proxy_url}


@dataclass(frozen=True)
class ItemRow:
    fetched_at_utc: str
    start: int
    name: str
    price_usd: str
    url: str


@dataclass
class Task:
    start: int
    tries_left: int


# -----------------------------
# Global rate limiter
# -----------------------------

class GlobalRateLimiter:
    """Гарантирует, что ВСЕ потоки вместе не сделают запросы чаще заданного интервала."""
    def __init__(self, min_interval_s: float):
        self.min_interval_s = min_interval_s
        self._lock = threading.Lock()
        self._next_time = 0.0

    def wait(self):
        with self._lock:
            now = time.time()
            if now < self._next_time:
                time.sleep(self._next_time - now)
            self._next_time = time.time() + self.min_interval_s


# -----------------------------
# Thread-safe proxy pool
# -----------------------------

class ProxyPool:
    """Выдаёт прокси строго по очереди, без пересечений между потоками."""
    def __init__(self, proxies: List[Proxy]):
        self._proxies = proxies
        self._lock = threading.Lock()
        self._idx = 0

    def acquire_next(self) -> Optional[Proxy]:
        with self._lock:
            if self._idx >= len(self._proxies):
                return None
            p = self._proxies[self._idx]
            self._idx += 1
            return p


# -----------------------------
# IO helpers
# -----------------------------

def load_proxies(path: str) -> List[Proxy]:
    proxies: List[Proxy] = []
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            parts = line.split(":")
            if len(parts) != 4:
                raise ValueError(f"Неверный формат прокси: {line} (ожидалось ip:port:user:pass)")
            host, port_s, user, password = parts
            proxies.append(Proxy(host=host, port=int(port_s), user=user, password=password))
    return proxies


def load_existing_index_from_xlsx(path: str) -> Tuple[Dict[int, Set[str]], Set[str]]:
    start_to_urls: Dict[int, Set[str]] = defaultdict(set)
    seen_urls: Set[str] = set()

    if not os.path.exists(path):
        return start_to_urls, seen_urls

    wb = load_workbook(path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 5:
            continue
        start = row[1]
        url = row[4]
        if isinstance(start, int) and isinstance(url, str) and url:
            start_to_urls[start].add(url)
            seen_urls.add(url)
    return start_to_urls, seen_urls


def ensure_workbook(out_path: str, append: bool) -> Tuple[Workbook, any]:
    if append and os.path.exists(out_path):
        wb = load_workbook(out_path)
        ws = wb.active
        if ws.max_row < 1:
            ws.append(["fetched_at_utc", "start", "name", "price_usd", "url"])
        return wb, ws

    wb = Workbook()
    ws = wb.active
    ws.title = "items"
    ws.append(["fetched_at_utc", "start", "name", "price_usd", "url"])
    return wb, ws


# -----------------------------
# FX (USD normalization)
# -----------------------------

def fetch_fx_rates_usd(session: requests.Session, timeout_s: int = 10) -> Dict[str, float]:
    """
    Frankfurter: base=USD => rates[EUR] = сколько EUR за 1 USD.
    EUR->USD: usd = eur / rates[EUR]
    """
    url = "https://api.frankfurter.dev/v1/latest"
    params = {"base": "USD", "symbols": "EUR,GBP,CHF"}
    r = session.get(url, params=params, timeout=timeout_s)
    r.raise_for_status()
    data = r.json()
    rates = data.get("rates", {})
    out: Dict[str, float] = {}
    for ccy in ("EUR", "GBP", "CHF"):
        v = rates.get(ccy)
        if isinstance(v, (int, float)) and v > 0:
            out[ccy] = float(v)
    return out


_money_num_re = re.compile(r"([0-9]+(?:[.,][0-9]+)?)")

def normalize_price_to_usd(price_raw: str, fx_usd_base: Dict[str, float]) -> str:
    if not price_raw:
        return ""
    s = price_raw.strip()
    if not s:
        return ""

    m = _money_num_re.search(s)
    if not m:
        return ""

    num_str = m.group(1).replace(",", ".")
    try:
        amount = float(num_str)
    except ValueError:
        return ""

    s_upper = s.upper()

    # USD
    if "$" in s or "USD" in s_upper:
        return f"${amount:.2f}"

    # EUR
    if "€" in s or "EUR" in s_upper:
        rate = fx_usd_base.get("EUR")
        return f"${(amount / rate):.2f}" if rate else ""

    # GBP
    if "£" in s or "GBP" in s_upper:
        rate = fx_usd_base.get("GBP")
        return f"${(amount / rate):.2f}" if rate else ""

    # CHF
    if "CHF" in s_upper or "FR" in s_upper:
        rate = fx_usd_base.get("CHF")
        return f"${(amount / rate):.2f}" if rate else ""

    # no currency => assume USD
    return f"${amount:.2f}"


# -----------------------------
# Steam parsing / requests
# -----------------------------

def steam_headers(appid: int) -> dict:
    return {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                      "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "X-Requested-With": "XMLHttpRequest",
        "Referer": f"https://steamcommunity.com/market/search?appid={appid}",
    }


def fetch_market_page(
    session: requests.Session,
    appid: int,
    start: int,
    count: int,
    proxy_dict: dict,
    timeout_s: int,
    sort_column: str,
    sort_dir: str,
) -> dict:
    url = "https://steamcommunity.com/market/search/render/"
    params = {
        "query": "",
        "start": start,
        "count": count,
        "search_descriptions": 0,
        "sort_column": sort_column,
        "sort_dir": sort_dir,
        "appid": appid,
    }
    r = session.get(url, params=params, headers=steam_headers(appid), proxies=proxy_dict, timeout=timeout_s)
    r.raise_for_status()
    return r.json()


def parse_items_from_results_html(results_html: str) -> List[Tuple[str, str, str]]:
    soup = BeautifulSoup(results_html, "lxml")
    rows = soup.select("a.market_listing_row_link")

    out: List[Tuple[str, str, str]] = []
    for a in rows:
        url = (a.get("href") or "").strip()
        name_el = a.select_one("span.market_listing_item_name")
        name = name_el.get_text(strip=True) if name_el else ""

        price_el = a.select_one("span.sale_price")
        if not price_el:
            price_el = a.select_one("span.normal_price[data-price]") or a.select_one("span.normal_price")
        price_raw = price_el.get_text(" ", strip=True) if price_el else ""

        if name and url:
            out.append((name, price_raw, url))
    return out


def expected_count_for_start(total_count: int, count: int, start: int) -> int:
    remaining = total_count - start
    if remaining <= 0:
        return 0
    return count if remaining >= count else remaining


def build_repair_starts(total_count: int, count: int, start_to_urls: Dict[int, Set[str]]) -> List[int]:
    starts = list(range(0, total_count, count))
    need: List[int] = []
    for s in starts:
        exp = expected_count_for_start(total_count, count, s)
        have = len(start_to_urls.get(s, set()))
        if have < exp:
            need.append(s)
    return need


def discover_total_count(
    proxies: List[Proxy],
    proxy_scheme: str,
    appid: int,
    timeout_s: int,
    sort_column: str,
    sort_dir: str,
) -> int:
    if not proxies:
        raise RuntimeError("Нет прокси для получения total_count.")
    session = requests.Session()
    p = proxies[0]
    payload = fetch_market_page(
        session=session,
        appid=appid,
        start=0,
        count=10,
        proxy_dict=p.as_requests_proxy(proxy_scheme),
        timeout_s=timeout_s,
        sort_column=sort_column,
        sort_dir=sort_dir,
    )
    tc = payload.get("total_count")
    if not isinstance(tc, int):
        raise RuntimeError("Не удалось прочитать total_count из ответа Steam.")
    return tc


# -----------------------------
# Writer thread
# -----------------------------

def excel_writer(
    out_path: str,
    append: bool,
    results_q: "queue.Queue[Optional[ItemRow]]",
    seen_urls: Set[str],
    save_every: int = 200,
):
    wb, ws = ensure_workbook(out_path, append)

    written = 0
    while True:
        item = results_q.get()
        if item is None:
            break

        if item.url in seen_urls:
            continue

        ws.append([item.fetched_at_utc, item.start, item.name, item.price_usd, item.url])
        seen_urls.add(item.url)
        written += 1

        if written % save_every == 0:
            wb.save(out_path)

    wb.save(out_path)


# -----------------------------
# Worker threads (queue + requeue)
# -----------------------------

def worker(
    worker_id: int,
    tasks_q: "queue.Queue[Task]",
    results_q: "queue.Queue[ItemRow]",
    proxy_pool: ProxyPool,
    appid: int,
    count: int,
    total_count: int,
    proxy_scheme: str,
    max_requests_per_proxy: int,
    timeout_s: int,
    base_delay_s: float,
    sort_column: str,
    sort_dir: str,
    fx_usd_base: Dict[str, float],
    rate_limiter: Optional[GlobalRateLimiter],
    stats: dict,
    stats_lock: threading.Lock,
):
    session = requests.Session()

    current_proxy: Optional[Proxy] = None
    current_proxy_left = 0

    def rotate_proxy() -> bool:
        nonlocal current_proxy, current_proxy_left
        current_proxy = proxy_pool.acquire_next()
        if current_proxy is None:
            return False
        current_proxy_left = max_requests_per_proxy
        return True

    if not rotate_proxy():
        return

    while True:
        try:
            task = tasks_q.get(timeout=2.0)
        except queue.Empty:
            return

        start = task.start
        exp = expected_count_for_start(total_count, count, start)
        if exp <= 0:
            tasks_q.task_done()
            continue

        ok = False
        last_err: Optional[str] = None
        status_code: Optional[int] = None

        # лимитер (глобально на все потоки)
        if rate_limiter:
            rate_limiter.wait()

        if current_proxy_left <= 0:
            if not rotate_proxy():
                tasks_q.task_done()
                return

        if current_proxy is None:
            tasks_q.task_done()
            return

        proxy_dict = current_proxy.as_requests_proxy(proxy_scheme)
        current_proxy_left -= 1

        try:
            payload = fetch_market_page(
                session=session,
                appid=appid,
                start=start,
                count=count,
                proxy_dict=proxy_dict,
                timeout_s=timeout_s,
                sort_column=sort_column,
                sort_dir=sort_dir,
            )

            if not payload.get("success", False):
                raise RuntimeError("success=false")

            items = parse_items_from_results_html(payload.get("results_html") or "")
            if len(items) < exp:
                raise RuntimeError(f"incomplete {len(items)}/{exp}")

            fetched_at = datetime.now(timezone.utc).isoformat()
            for name, price_raw, url in items:
                price_usd = normalize_price_to_usd(price_raw, fx_usd_base)
                results_q.put(ItemRow(fetched_at, start, name, price_usd, url))

            ok = True

        except requests.HTTPError as e:
            status_code = getattr(e.response, "status_code", None)
            last_err = f"HTTP {status_code}"
        except Exception as e:
            last_err = str(e)

        # статистика
        with stats_lock:
            stats["processed"] += 1
            if ok:
                stats["ok"] += 1
            else:
                stats["fail"] += 1
                if status_code == 429:
                    stats["429"] += 1
                elif status_code == 403:
                    stats["403"] += 1
                else:
                    stats["other_err"] += 1

        # если не ок — возвращаем в очередь (если есть попытки)
        if not ok:
            # Для 403/429 лучше сразу сменить прокси
            if status_code in (403, 429):
                rotate_proxy()

            if task.tries_left > 0:
                # backoff перед повтором (чем меньше tries_left — тем больше пауза)
                backoff = min(20.0, base_delay_s * (2.0 ** (max(0, 5 - task.tries_left)))) + random.uniform(0, 1.0)
                time.sleep(backoff)
                tasks_q.put(Task(start=start, tries_left=task.tries_left - 1))
            # иначе просто сдаёмся по этой странице

        tasks_q.task_done()

        time.sleep(base_delay_s + random.uniform(0, 0.4))


# -----------------------------
# Main
# -----------------------------

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--proxies", required=True, help="Путь к proxies.txt (ip:port:user:pass)")
    ap.add_argument("--out", default="steam_items.xlsx", help="Выходной Excel файл")
    ap.add_argument("--append", action="store_true", help="Дописать в существующий Excel")
    ap.add_argument("--repair", action="store_true", help="Докачать страницы, где записано меньше ожидаемого")

    ap.add_argument("--appid", type=int, default=252490, help="Steam appid")
    ap.add_argument("--count", type=int, default=10, help="Сколько предметов за запрос")
    ap.add_argument("--workers", type=int, default=5, help="Количество потоков")

    ap.add_argument("--proxy-scheme", default="http", help="http или socks5")
    ap.add_argument("--max-req-per-proxy", type=int, default=10, help="Максимум запросов на один прокси")
    ap.add_argument("--timeout", type=int, default=25, help="Timeout запроса, секунд")
    ap.add_argument("--delay", type=float, default=0.4, help="Базовая пауза между итерациями воркера")

    ap.add_argument("--sort-column", default="name", help="name | price | quantity")
    ap.add_argument("--sort-dir", default="desc", help="asc | desc")

    ap.add_argument("--total-count", type=int, default=0, help="Если знаете total_count (например 5026) — задайте")
    ap.add_argument("--max-items", type=int, default=0, help="Ограничить общее число предметов (тесты). 0=нет")

    # Сколько раз возвращать задачу в очередь при ошибке
    ap.add_argument("--task-retries", type=int, default=6, help="Сколько раз возвращать start в очередь при падении")

    # Глобальный интервал между запросами всех потоков (главная защита от 429)
    ap.add_argument("--global-interval", type=float, default=0.9,
                    help="Глобальный интервал между любыми запросами (сек). 0=выкл")

    args = ap.parse_args()

    proxies = load_proxies(args.proxies)
    if not proxies:
        raise SystemExit("Список прокси пустой.")

    # FX rates (1 раз)
    fx_usd_base: Dict[str, float] = {}
    try:
        fx_usd_base = fetch_fx_rates_usd(requests.Session(), timeout_s=10)
    except Exception:
        fx_usd_base = {}

    # Индекс из Excel
    start_to_urls: Dict[int, Set[str]] = defaultdict(set)
    seen_urls: Set[str] = set()
    if args.append:
        start_to_urls, seen_urls = load_existing_index_from_xlsx(args.out)

    # total_count
    if args.total_count > 0:
        total_count = args.total_count
    else:
        total_count = discover_total_count(
            proxies=proxies,
            proxy_scheme=args.proxy_scheme,
            appid=args.appid,
            timeout_s=args.timeout,
            sort_column=args.sort_column,
            sort_dir=args.sort_dir,
        )

    if args.max_items and args.max_items > 0:
        total_count = min(total_count, args.max_items)

    # Подготовка starts
    if args.repair:
        starts = build_repair_starts(total_count, args.count, start_to_urls)
    else:
        starts_all = list(range(0, total_count, args.count))
        starts = []
        for s in starts_all:
            exp = expected_count_for_start(total_count, args.count, s)
            have = len(start_to_urls.get(s, set())) if args.append else 0
            if args.append and have >= exp:
                continue
            starts.append(s)

    # Очередь задач
    tasks_q: "queue.Queue[Task]" = queue.Queue()
    for s in starts:
        tasks_q.put(Task(start=s, tries_left=args.task_retries))

    capacity_pages = len(proxies) * args.max_req_per_proxy
    print(f"Прокси: {len(proxies)} шт.")
    print(f"Сортировка: sort_column={args.sort_column}, sort_dir={args.sort_dir}")
    print(f"Лимит: {args.max_req_per_proxy} запрос(а) на прокси -> ёмкость {capacity_pages} страниц (на один проход)")
    print(f"total_count (учтённый): {total_count} предметов -> {((total_count + args.count - 1) // args.count)} страниц")
    print(f"Потоки: {args.workers}")
    print(f"Global interval: {args.global_interval} сек")
    if fx_usd_base:
        print(f"FX: загружены курсы для {', '.join(sorted(fx_usd_base.keys()))} (конвертация в USD включена)")
    else:
        print("FX: не удалось загрузить курсы (конвертация отключена)")
    if args.append:
        print(f"Append-режим: url в файле={len(seen_urls)}, страниц(start) в файле={len(start_to_urls)}")
    print(f"{'Repair' if args.repair else 'Run'}: стартовых страниц в очереди={len(starts)}")

    if tasks_q.qsize() == 0:
        print("Нечего делать.")
        return

    proxy_pool = ProxyPool(proxies)
    results_q: "queue.Queue[Optional[ItemRow]]" = queue.Queue(maxsize=12000)

    rate_limiter = GlobalRateLimiter(args.global_interval) if args.global_interval and args.global_interval > 0 else None

    # Stats
    stats = {"processed": 0, "ok": 0, "fail": 0, "429": 0, "403": 0, "other_err": 0}
    stats_lock = threading.Lock()

    writer_t = threading.Thread(target=excel_writer, args=(args.out, args.append, results_q, seen_urls), daemon=True)
    writer_t.start()

    threads = []
    for i in range(args.workers):
        t = threading.Thread(
            target=worker,
            args=(
                i,
                tasks_q,
                results_q,
                proxy_pool,
                args.appid,
                args.count,
                total_count,
                args.proxy_scheme,
                args.max_req_per_proxy,
                args.timeout,
                args.delay,
                args.sort_column,
                args.sort_dir,
                fx_usd_base,
                rate_limiter,
                stats,
                stats_lock,
            ),
            daemon=True,
        )
        t.start()
        threads.append(t)

    # Ждём пока очередь задач будет полностью обработана (включая повторно добавленные)
    tasks_q.join()

    # Останавливаем writer
    results_q.put(None)
    writer_t.join()

    with stats_lock:
        print("Статистика:", stats)
    print(f"Готово. Файл: {args.out}")


if __name__ == "__main__":
    main()
