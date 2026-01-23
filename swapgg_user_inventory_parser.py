#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
swapgg_user_inventory_parser.py

Selenium + Selenium-Wire парсер собственного инвентаря swap.gg (USER inventory) для Rust.
Логика:
  1) Открываем https://swap.gg/ru/trade
  2) Ждём пока пользователь вручную авторизуется в Steam/swap.gg (в браузере)
  3) Спрашиваем подтверждение (y/n)
  4) Переключаем ЛЕВУЮ панель (инвентарь пользователя) на Rust
  5) Скроллим левый recycle-scroller
  6) Забираем данные из Network ответов POST https://swap.gg/api/inventory/user
  7) Фильтруем: product.category == "RUST" и product.conditions == [] (пусто)
  8) Сохраняем в Excel: fetched_at_utc, idx, name, price_usd, count, stack, item_ids

Зависимости:
  pip install selenium selenium-wire openpyxl
  pip uninstall blinker -y
  pip install blinker==1.6.2
"""

import argparse
import json
import time
from dataclasses import dataclass
from datetime import datetime, timezone
from typing import Dict, List, Optional, Set, Tuple

from openpyxl import Workbook, load_workbook

from seleniumwire import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


# -----------------------------
# Data model
# -----------------------------

@dataclass(frozen=True)
class ItemRow:
    fetched_at_utc: str
    idx: int
    name: str
    price_usd: str
    count: int
    stack: str
    item_ids: str


# -----------------------------
# Excel helpers
# -----------------------------

def ensure_workbook(out_path: str, append: bool):
    if append:
        try:
            wb = load_workbook(out_path)
            ws = wb.active
            if ws.max_row < 1:
                ws.append(["fetched_at_utc", "idx", "name", "price_usd", "count", "stack", "item_ids"])
            return wb, ws
        except Exception:
            pass

    wb = Workbook()
    ws = wb.active
    ws.title = "items"
    ws.append(["fetched_at_utc", "idx", "name", "price_usd", "count", "stack", "item_ids"])
    return wb, ws


def save_rows_to_xlsx(out_path: str, rows: List[ItemRow], append: bool, save_every: int = 200, verbose: bool = False):
    wb, ws = ensure_workbook(out_path, append)
    written = 0
    for r in rows:
        ws.append([r.fetched_at_utc, r.idx, r.name, r.price_usd, r.count, r.stack, r.item_ids])
        written += 1
        if written % save_every == 0:
            wb.save(out_path)
            if verbose:
                print(f"[xlsx] saved {written} rows -> {out_path}")
    wb.save(out_path)
    if verbose:
        print(f"[xlsx] final save {written} rows -> {out_path}")


# -----------------------------
# Helpers: page / ui
# -----------------------------

def looks_like_cloudflare_challenge(driver) -> bool:
    title = (driver.title or "").lower()
    src = (driver.page_source or "").lower()
    return ("just a moment" in title) or ("checking your browser" in src)


def accept_cookies_if_present(driver, verbose: bool = False):
    # максимально простые эвристики
    candidates_xpath = [
        "//button[contains(translate(., 'ACEGPT', 'acegpt'), 'accept')]",
        "//button[contains(translate(., 'ACEGPT', 'acegpt'), 'agree')]",
        "//button[contains(., 'Принять')]",
        "//button[contains(., 'Соглас')]",
        "//button[contains(., 'Ок')]",
        "//button[contains(., 'OK')]",
    ]
    for xp in candidates_xpath:
        try:
            btn = WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.XPATH, xp)))
            driver.execute_script("arguments[0].click();", btn)
            if verbose:
                print("[info] Нажал кнопку принятия cookies.")
            time.sleep(0.5)
            return
        except Exception:
            pass


def wait_until_trade_ready(driver, timeout_s: int = 60, verbose: bool = False):
    t0 = time.time()
    while time.time() - t0 < timeout_s:
        if looks_like_cloudflare_challenge(driver):
            if verbose:
                print("[info] Cloudflare challenge (Just a moment...). Если появится проверка — пройди её вручную.")
            time.sleep(2.0)
            continue
        try:
            WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, ".vue-recycle-scroller, button[title]"))
            )
            return
        except Exception:
            time.sleep(1.0)

    if verbose:
        print("[warn] Таймаут ожидания готовности trade-страницы.")


def _safe_click(driver, el, verbose: bool = False) -> bool:
    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
        time.sleep(0.15)
    except Exception:
        pass
    try:
        el.click()
        return True
    except Exception:
        pass
    try:
        driver.execute_script("arguments[0].click();", el)
        return True
    except Exception as e:
        if verbose:
            print(f"[warn] click failed: {e}")
        return False


def find_game_buttons_rust(driver) -> List:
    # чаще всего есть title="Rust"
    btns = []
    try:
        btns = driver.find_elements(By.CSS_SELECTOR, 'button[title="Rust"]')
    except Exception:
        btns = []
    if btns:
        return btns
    # фолбэк по тексту
    try:
        return driver.find_elements(By.XPATH, "//button[contains(., 'Rust')]")
    except Exception:
        return []


def click_rust_left_panel(driver, verbose: bool = False) -> bool:
    """
    На странице обычно 2 кнопки Rust:
      - левая: USER inventory
      - правая: SITE inventory
    Нам нужна ЛЕВАЯ: берём кнопку с минимальным X (самая левая).
    """
    btns = find_game_buttons_rust(driver)
    if not btns:
        if verbose:
            print("[warn] Не нашёл кнопки Rust.")
        return False

    best = None
    best_x = 10**9
    for b in btns:
        try:
            x = driver.execute_script("return arguments[0].getBoundingClientRect().x;", b)
            if x is not None and x < best_x:
                best_x = x
                best = b
        except Exception:
            continue

    if best is None:
        best = btns[0]

    ok = _safe_click(driver, best, verbose=verbose)
    if verbose:
        print(f"[info] Rust(LEFT) click_ok={ok} x={best_x:.1f} (user inventory)")
    time.sleep(1.0)
    return ok


def find_recycle_scroller_left(driver, wait_s: int = 25, verbose: bool = False):
    """
    Ищем vue-recycle-scroller ЛЕВОЙ колонки (инвентарь пользователя).
    Выбираем элемент с минимальной координатой X.
    """
    wait = WebDriverWait(driver, wait_s)

    try:
        wait.until(
            EC.presence_of_element_located((By.CSS_SELECTOR, ".vue-recycle-scroller"))
        )
    except Exception:
        if verbose:
            print("[warn] Не найден ни один .vue-recycle-scroller")
        return None

    scrollers = driver.find_elements(By.CSS_SELECTOR, ".vue-recycle-scroller")
    if not scrollers:
        if verbose:
            print("[warn] Список scroller'ов пуст.")
        return None

    best = None
    best_x = 10**9  # ищем минимальный X (самый левый)

    for el in scrollers:
        try:
            rect = driver.execute_script(
                "return arguments[0].getBoundingClientRect();", el
            )
            x = rect.get("x", None)
            w = rect.get("width", None)

            if verbose:
                print(f"[debug] scroller rect x={x:.1f} w={w:.1f}")

            if x is not None and x < best_x:
                best_x = x
                best = el
        except Exception:
            continue

    if best and verbose:
        print(f"[info] Выбран ЛЕВЫЙ scroller (x={best_x:.1f})")

    return best



# -----------------------------
# Network parsing: /api/inventory/user
# -----------------------------

def price_int_to_usd(price_int: int) -> str:
    # 75417 -> 754.17
    try:
        return f"${(price_int / 100.0):.2f}"
    except Exception:
        return ""


def harvest_user_inventory_from_requests(driver, verbose: bool = False) -> List[dict]:
    """
    Забираем JSON-ответы из selenium-wire запросов к /api/inventory/user
    Возвращаем плоский список items (как в data.items).
    """
    out: List[dict] = []
    try:
        reqs = list(getattr(driver, "requests", []))
    except Exception:
        reqs = []

    for r in reqs:
        try:
            if not r.response:
                continue
            url = (r.url or "")
            if "/api/inventory/user" not in url:
                continue
            # ожидаем 201 Created
            if int(getattr(r.response, "status_code", 0)) not in (200, 201):
                continue

            body = r.response.body
            if not body:
                continue

            # body может быть bytes
            if isinstance(body, (bytes, bytearray)):
                txt = body.decode("utf-8", errors="ignore")
            else:
                txt = str(body)

            data = json.loads(txt)
            items = (data.get("data") or {}).get("items") or []
            if isinstance(items, list):
                out.extend(items)

        except Exception as e:
            if verbose:
                print(f"[debug] harvest parse fail: {e}")
            continue

    # чистим буфер, чтобы не разрастался
    try:
        driver.requests.clear()
    except Exception:
        pass

    return out


def filter_and_convert_items(raw_items: List[dict], verbose: bool = False) -> List[Tuple[str, str, int, str, str]]:
    """
    Фильтр:
      - product.category == "RUST"
      - product.conditions == [] (пусто)  -> иначе НЕ добавляем
    Выход: (name, price_usd, count, stack, item_ids_csv)
    """
    out: List[Tuple[str, str, int, str, str]] = []
    for it in raw_items:
        try:
            product = it.get("product") or {}
            category = product.get("category")
            if category != "RUST":
                continue

            conditions = product.get("conditions") or []
            # если есть условия — предмет "плохой" (не добавляем)
            if isinstance(conditions, list) and len(conditions) > 0:
                continue

            name = product.get("name") or ""
            if not name:
                name = (product.get("metadata") or {}).get("itemName") or ""

            price_int = it.get("price")
            if not isinstance(price_int, int):
                continue
            price_usd = price_int_to_usd(price_int)

            count = it.get("count")
            if not isinstance(count, int):
                count = 1

            stack = it.get("stack") or ""
            item_ids = it.get("itemIds") or []
            if isinstance(item_ids, list):
                item_ids_csv = ",".join([str(x) for x in item_ids])
            else:
                item_ids_csv = ""

            if name:
                out.append((name, price_usd, count, stack, item_ids_csv))

        except Exception as e:
            if verbose:
                print(f"[debug] filter fail: {e}")
            continue
    return out


# -----------------------------
# Scrolling loop
# -----------------------------

def scroll_and_collect_user_inventory(
    driver,
    scroller_el,
    max_items: int,
    pause_s: float,
    stable_rounds: int,
    verbose: bool,
) -> List[ItemRow]:
    """
    Скроллим ЛЕВЫЙ скроллер.
    На каждом шаге забираем новые ответы /api/inventory/user из Network.
    Останавливаемся когда:
      - собрано max_items (если >0)
      - stable_rounds шагов подряд не прилетало новых уникальных предметов
    """
    seen_keys: Set[str] = set()
    rows: List[ItemRow] = []
    idx = 0
    stable = 0

    def add_items(converted: List[Tuple[str, str, int, str, str]]):
        nonlocal idx
        now = datetime.now(timezone.utc).isoformat()
        added = 0
        for name, price_usd, count, stack, item_ids_csv in converted:
            # ключ лучше делать по stack (уникально для предмета/стака)
            key = stack or (item_ids_csv.split(",")[0] if item_ids_csv else f"{name}|{price_usd}|{count}")
            if key in seen_keys:
                continue
            seen_keys.add(key)
            rows.append(ItemRow(now, idx, name, price_usd, count, stack, item_ids_csv))
            idx += 1
            added += 1
        return added

    # небольшой прогрев: подождать первые запросы
    time.sleep(1.0)
    raw = harvest_user_inventory_from_requests(driver, verbose=verbose)
    added0 = add_items(filter_and_convert_items(raw, verbose=verbose))
    if verbose:
        print(f"[info] start collected={len(rows)} (+{added0})")

    while True:
        if max_items > 0 and len(rows) >= max_items:
            if verbose:
                print(f"[info] reached max-items={max_items}")
            break

        # скролл вниз быстрее (0.95 экрана)
        try:
            driver.execute_script(
                "arguments[0].scrollTop = arguments[0].scrollTop + Math.floor(arguments[0].clientHeight * 0.95);",
                scroller_el
            )
        except Exception:
            try:
                driver.execute_script("arguments[0].scrollTop = arguments[0].scrollHeight;", scroller_el)
            except Exception:
                pass

        time.sleep(pause_s)

        raw_items = harvest_user_inventory_from_requests(driver, verbose=verbose)
        converted = filter_and_convert_items(raw_items, verbose=verbose)
        added = add_items(converted)

        if verbose:
            print(f"[info] collected={len(rows)} (+{added}) stable={stable}/{stable_rounds}")

        if added == 0:
            stable += 1
        else:
            stable = 0

        if stable >= stable_rounds:
            if verbose:
                print("[info] list seems stable (no new items). stopping.")
            break

    return rows


# -----------------------------
# WebDriver setup
# -----------------------------

def build_chrome_driver(headless: bool, user_data_dir: Optional[str], verbose: bool):
    opts = ChromeOptions()
    if headless:
        opts.add_argument("--headless=new")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--window-size=1500,900")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--lang=ru-RU")
    opts.add_argument("--disable-blink-features=AutomationControlled")

    if user_data_dir:
        opts.add_argument(f"--user-data-dir={user_data_dir}")
        if verbose:
            print(f"[info] using user-data-dir: {user_data_dir}")

    sw_opts = {
        # можно выключить verify_ssl, иногда помогает с некоторыми окружениями
        "verify_ssl": False,
        # ограничим хранение (не всегда работает во всех версиях)
        # "request_storage": "memory",
    }

    driver = webdriver.Chrome(options=opts, seleniumwire_options=sw_opts)
    try:
        driver.scopes = [".*swap\\.gg/api/inventory/user.*"]  # перехватываем только нужное
    except Exception:
        pass
    return driver


# -----------------------------
# Main
# -----------------------------

def main():
    ap = argparse.ArgumentParser()

    ap.add_argument("--out", default="swap_user_rust.xlsx", help="Выходной Excel файл")
    ap.add_argument("--append", action="store_true", help="Дописать в существующий Excel")
    ap.add_argument("--url", default="https://swap.gg/ru/trade", help="Страница trade")
    ap.add_argument("--max-items", type=int, default=1000, help="Сколько максимум собрать (0=без лимита)")
    ap.add_argument("--scroll-pause", type=float, default=0.45, help="Пауза между скроллами (быстрее = меньше)")
    ap.add_argument("--stable-rounds", type=int, default=25, help="Остановка если N шагов подряд нет новых предметов")
    ap.add_argument("--challenge-wait", type=int, default=120, help="Сколько секунд ждать ручной CF-проверки")
    ap.add_argument("--headless", action="store_true", help="Headless режим (лучше выключить)")
    ap.add_argument("--user-data-dir", default="", help="Путь к профилю Chrome (чтобы сохранять сессию). Например: ./chrome_profile_user")
    ap.add_argument("--verbose", action="store_true", help="Подробные логи")

    args = ap.parse_args()
    user_data_dir = args.user_data_dir.strip() or None

    driver = build_chrome_driver(headless=args.headless, user_data_dir=user_data_dir, verbose=args.verbose)

    try:
        if args.verbose:
            print(f"[info] open: {args.url}")
        driver.get(args.url)

        wait_until_trade_ready(driver, timeout_s=args.challenge_wait, verbose=args.verbose)
        accept_cookies_if_present(driver, verbose=args.verbose)

        if looks_like_cloudflare_challenge(driver):
            print("[warn] Cloudflare challenge всё ещё виден. Пройди вручную в окне браузера.")
            wait_until_trade_ready(driver, timeout_s=args.challenge_wait, verbose=args.verbose)

        # ---- ручной логин ----
        print("\n[login] Окно браузера открыто.")
        print("[login] Авторизуйся на swap.gg через Steam (вручную).")
        print("[login] Когда закончишь — вернись в консоль.\n")

        while True:
            ans = input("Вы авторизовались в Steam? (y/n): ").strip().lower()
            if ans == "y":
                break
            print("Ок, жду. Авторизуйся в окне браузера и снова введи y.\n")

        # ---- переключаем USER inventory на Rust (левая кнопка) ----
        if not click_rust_left_panel(driver, verbose=args.verbose):
            print("[error] Не удалось переключить Rust на левой панели. Попробуй вручную выбрать Rust и перезапустить.")
            # не выходим сразу, вдруг уже выбран
            time.sleep(1.0)

        # ---- находим ЛЕВЫЙ scroller ----
        scroller = find_recycle_scroller_left(driver, wait_s=25, verbose=args.verbose)

        if scroller is None:
            print("[error] Не удалось найти scroller инвентаря пользователя (левая колонка).")
            return


        # ---- сбор ----
        rows = scroll_and_collect_user_inventory(
            driver=driver,
            scroller_el=scroller,
            max_items=args.max_items,
            pause_s=args.scroll_pause,
            stable_rounds=args.stable_rounds,
            verbose=args.verbose,
        )

        if args.verbose:
            print(f"[info] total collected: {len(rows)}")

        save_rows_to_xlsx(args.out, rows, append=args.append, save_every=200, verbose=args.verbose)
        print(f"Готово. Файл: {args.out}")

    finally:
        try:
            driver.quit()
        except Exception:
            pass


if __name__ == "__main__":
    main()
