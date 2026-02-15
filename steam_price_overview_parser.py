import argparse
import os
import queue
import random
import re
import threading
import time
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from typing import Dict, List, Optional, Set, Tuple
from urllib.parse import quote, quote_plus

import requests
from openpyxl import Workbook, load_workbook


# -----------------------------
# Data models
# -----------------------------

@dataclass(frozen=True)
class Proxy:
    host: str
    port: int
    user: str
    password: str

    def as_requests_proxy(self, scheme: str = "http") -> dict:
        u = quote(self.user, safe="")
        p = quote(self.password, safe="")
        proxy_url = f"{scheme}://{u}:{p}@{self.host}:{self.port}"
        return {"http": proxy_url, "https": proxy_url}


@dataclass(frozen=True)
class PriceRow:
    fetched_at_utc: str
    name: str # Оригинальное имя
    price_usd: str
    url: str


@dataclass
class Task:
    name: str
    tries_left: int


# -----------------------------
# Global rate limiter
# -----------------------------

class GlobalRateLimiter:
    """Гарантирует, что ВСЕ потоки вместе не сделают запросы чаще заданного интервала."""
    def __init__(self, min_interval_s: float):
        self.min_interval_s = min_interval_s
        self._lock = threading.Lock()
        self._next_time = 0.0

    def wait(self):
        with self._lock:
            now = time.time()
            if now < self._next_time:
                time.sleep(self._next_time - now)
            self._next_time = time.time() + self.min_interval_s


# -----------------------------
# Thread-safe proxy pool
# -----------------------------

class ProxyPool:
    """
    Выдаёт прокси по очереди, без пересечений между потоками.
    Если прокси закончились — начинает заново (pass++),
    но не более max_passes раз.
    """
    def __init__(self, proxies: List[Proxy], max_passes: int = 1):
        self._proxies = proxies
        self._lock = threading.Lock()
        self._idx = 0
        self._pass = 1
        self._max_passes = max(1, int(max_passes))

    def acquire_next(self) -> Optional[Proxy]:
        with self._lock:
            if not self._proxies:
                return None

            if self._idx >= len(self._proxies):
                # закончили текущий проход по списку прокси
                self._pass += 1
                if self._pass > self._max_passes:
                    return None
                self._idx = 0  # новый проход

            p = self._proxies[self._idx]
            self._idx += 1
            return p

    def state(self) -> Tuple[int, int]:
        """(текущий проход, максимум проходов)"""
        with self._lock:
            return self._pass, self._max_passes


# -----------------------------
# IO helpers
# -----------------------------

def load_proxies(path: str) -> List[Proxy]:
    proxies: List[Proxy] = []
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            parts = line.split(":")
            if len(parts) != 4:
                raise ValueError(f"Неверный формат прокси: {line} (ожидалось ip:port:user:pass)")
            host, port_s, user, password = parts
            proxies.append(Proxy(host=host, port=int(port_s), user=user, password=password))
    return proxies


def load_names(path: str) -> List[str]:
    names: List[str] = []
    if not os.path.exists(path):
        return names
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                names.append(line)
    return names



def load_processed_names(out_path: str) -> Set[str]:
    processed = set()
    if not os.path.exists(out_path):
        return processed
    try:
        wb = load_workbook(out_path)
        ws = wb.active
        # Предполагаем, что имя во 2-й колонке (индекс 1, если 0-based)
        # rows generator returns tuples
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= 2:
                name = row[1]
                if name:
                    processed.add(str(name))
    except Exception as e:
        print(f"Ошибка чтения {out_path} для resume: {e}")
    return processed


def ensure_workbook(out_path: str, append: bool) -> Tuple[Workbook, any]:
    if os.path.exists(out_path):
        try:
            wb = load_workbook(out_path)
            ws = wb.active
            if ws.max_row < 1:
                 ws.append(["fetched_at_utc", "name", "price_usd", "url"])
            return wb, ws
        except Exception as e:
            print(f"Не удалось открыть {out_path}, создаем новый. Ошибка: {e}")

    wb = Workbook()
    ws = wb.active
    ws.title = "prices"
    ws.append(["fetched_at_utc", "name", "price_usd", "url"])
    return wb, ws


# -----------------------------
# Price parsing / requests
# -----------------------------

def steam_headers() -> dict:
    return {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                      "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "X-Requested-With": "XMLHttpRequest",
        "Referer": "https://steamcommunity.com/market/",
    }


def fetch_price_overview(
    session: requests.Session,
    appid: int,
    market_hash_name: str,
    currency: int,
    proxy_dict: dict,
    timeout_s: int,
) -> dict:
    url = "https://steamcommunity.com/market/priceoverview/"
    params = {
        "country": "US",
        "currency": currency,
        "appid": appid,
        "market_hash_name": market_hash_name,
    }
    r = session.get(url, params=params, headers=steam_headers(), proxies=proxy_dict, timeout=timeout_s)
    r.raise_for_status()
    # Иногда Steam возвращает null, если предмета нет или ошибка
    if not r.content:
        raise RuntimeError("Empty response content")
    return r.json()


_money_re = re.compile(r"([0-9]+(?:[.,][0-9]+)?)")

def parse_price_value(price_str: str) -> str:
    """
    Принимает строку вида "$2.47" или "247,54 руб."
    Возвращает чистую строку цены "2.47" или "247.54" (с точкой).
    Если не удалось распарсить - возвращает исходную или пустую строку.
    В данном скрипте мы запрашиваем в USD (currency=1), ожидаем формат "$X.XX".
    """
    if not price_str:
        return ""
    
    # Пытаемся найти число
    m = _money_re.search(price_str)
    if not m:
        return ""
    
    val = m.group(1).replace(",", ".")
    return val

# -----------------------------
# Writer thread
# -----------------------------

def _save_workbook_safe(wb: Workbook, path: str, retries: int = 5):
    for i in range(retries):
        try:
            wb.save(path)
            return
        except PermissionError:
            print(f"[Writer] Ошибка доступа к файлу {path} (возможно открыт?). Попытка {i+1}/{retries} через 5 сек...")
            time.sleep(8)
        except Exception as e:
            print(f"[Writer] Ошибка сохранения {path}: {e}")
            return
    print(f"[Writer] Не удалось сохранить {path} после {retries} попыток. Данные могут быть потеряны.")


def excel_writer(
    out_path: str,
    append: bool,
    results_q: "queue.Queue[Optional[PriceRow]]",
    save_every: int = 50,
):
    # append по сути регулируется тем, передали ли мы append=True В main
    # но тут мы уже внутри ensure_workbook проверяем наличие файла
    wb, ws = ensure_workbook(out_path, append)

    written = 0
    while True:
        item = results_q.get()
        if item is None:
            break

        ws.append([item.fetched_at_utc, item.name, item.price_usd, item.url])
        written += 1

        if written % save_every == 0:
            _save_workbook_safe(wb, out_path)

    _save_workbook_safe(wb, out_path)


# -----------------------------
# Worker threads
# -----------------------------

def worker(
    worker_id: int,
    tasks_q: "queue.Queue[Task]",
    results_q: "queue.Queue[PriceRow]",
    proxy_pool: ProxyPool,
    appid: int,
    currency: int,
    proxy_scheme: str,
    max_requests_per_proxy: int,
    timeout_s: int,
    base_delay_s: float,
    rate_limiter: Optional[GlobalRateLimiter],
    stats: dict,
    stats_lock: threading.Lock,
    verbose: bool,
    stop_evt: threading.Event,
):
    session = requests.Session()

    current_proxy: Optional[Proxy] = None
    current_proxy_left = 0

    def rotate_proxy() -> bool:
        nonlocal current_proxy, current_proxy_left
        current_proxy = proxy_pool.acquire_next()
        if current_proxy is None:
            # прокси/проходы закончились для этого потока
            # НЕ вызываем stop_evt.set(), чтобы не мешать другим потокам дорабатывать свои прокси
            return False
        current_proxy_left = max_requests_per_proxy
        return True

    if not rotate_proxy():
        if verbose:
            print(f"[worker {worker_id}] No proxies available (init). Exiting.")
        return

    while True:
        if stop_evt.is_set():
            return
        try:
            task = tasks_q.get(timeout=2.0)
        except queue.Empty:
            return

        name = task.name
        if verbose:
            print(f"[worker {worker_id}] Processing: {name}")

        ok = False
        last_err: Optional[str] = None
        status_code: Optional[int] = None

        # лимитер
        if rate_limiter:
            rate_limiter.wait()

        if current_proxy_left <= 0:
            if not rotate_proxy():
                tasks_q.task_done()
                return

        if current_proxy is None:
            tasks_q.task_done()
            return

        proxy_dict = current_proxy.as_requests_proxy(proxy_scheme)
        current_proxy_left -= 1

        try:
            payload = fetch_price_overview(
                session=session,
                appid=appid,
                market_hash_name=name,
                currency=currency,
                proxy_dict=proxy_dict,
                timeout_s=timeout_s,
            )

            # success может быть false, но если это просто "нет лотов", то не надо ретраить как ошибку сети
            # Обычно если нет лотов, success=true, но lowest_price нет.
            
            lowest_price_raw = payload.get("lowest_price", "")
            median_price_raw = payload.get("median_price", "")
            price_raw = lowest_price_raw if lowest_price_raw else median_price_raw
            price_val = parse_price_value(price_raw)
            
            fetched_at = datetime.now(timezone.utc).isoformat()
            # item URL
            url_name = quote(name)
            item_url = f"https://steamcommunity.com/market/listings/{appid}/{url_name}"

            if price_val:
                price_final = f"${price_val}"
            else:
                price_final = "" # Пустая цена - тоже результат (предмета нет)

            results_q.put(PriceRow(fetched_at, name, price_final, item_url))

            ok = True
            if verbose:
                print(f"[worker {worker_id}] OK: {name} -> {price_final}")

        except requests.HTTPError as e:
            status_code = getattr(e.response, "status_code", None)
            last_err = f"HTTP {status_code}"
            # 500/502/... - ретраим. 404 - странно для priceoverview, но обработаем.
        except Exception as e:
            last_err = str(e)

        # Статистика
        with stats_lock:
            stats["processed"] += 1
            if ok:
                stats["ok"] += 1
            else:
                stats["fail"] += 1
                if status_code == 429:
                    stats["429"] += 1
                elif status_code == 403:
                    stats["403"] += 1
                else:
                    stats["other_err"] += 1

        # Обработка ошибок / повторов
        if not ok:
            if verbose:
                sc = status_code if status_code is not None else "-"
                print(f"[worker {worker_id}] FAIL {sc} err={last_err} Item: {name}")

            if status_code in (403, 429):
                rotate_proxy()

            if task.tries_left > 0:
                # backoff
                backoff = min(20.0, base_delay_s * (2.0 ** (max(0, 3 - task.tries_left)))) + random.uniform(0, 1.0)
                time.sleep(backoff)
                tasks_q.put(Task(name=name, tries_left=task.tries_left - 1))
            else:
                # Все попытки исчерпаны - пишем пустую строку, чтобы не потерять, что мы пытались
                url_name = quote(name)
                item_url = f"https://steamcommunity.com/market/listings/{appid}/{url_name}"
                fetched_at = datetime.now(timezone.utc).isoformat()
                results_q.put(PriceRow(fetched_at, name, "ERROR", item_url))
                print(f"[worker {worker_id}] DROP {name} after retries (saved as ERROR)")

        tasks_q.task_done()
        time.sleep(base_delay_s + random.uniform(0, 0.4))


def progress_printer(progress_every_s: float, tasks_q: "queue.Queue[Task]", stats: dict, lock: threading.Lock, stop_evt: threading.Event):
    if progress_every_s <= 0:
        return
    while not stop_evt.is_set():
        time.sleep(progress_every_s)
        with lock:
            s = dict(stats)
        try:
            qn = tasks_q.qsize()
        except:
            qn = -1
        print(f"[progress] queue={qn} processed={s.get('processed',0)} ok={s.get('ok',0)} "
              f"fail={s.get('fail',0)} 429={s.get('429',0)}")


# -----------------------------
# Main
# -----------------------------

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--names", default="names.txt", help="Файл с названиями предметов (построчно)")
    ap.add_argument("--proxies", default="proxies.txt", help="Путь к proxies.txt (ip:port:user:pass)")
    ap.add_argument("--out", default="steam_prices.xlsx", help="Выходной Excel файл")
    ap.add_argument("--append", action="store_true", help="Дописать в существующий Excel (автоматически включается, если файл есть)")

    ap.add_argument("--appid", type=int, default=252490, help="Steam appid (default Rust 252490)")
    ap.add_argument("--currency", type=int, default=1, help="Currency ID (1=USD, 5=RUB). Default 1 (USD)")
    
    ap.add_argument("--workers", type=int, default=1, help="Количество потоков")
    ap.add_argument("--proxy-scheme", default="http", help="http или socks5")
    ap.add_argument("--max-req-per-proxy", type=int, default=50, help="Максимум запросов на один прокси")
    ap.add_argument("--timeout", type=int, default=25, help="Timeout запроса, секунд")
    ap.add_argument("--delay", type=float, default=1.5, help="Базовая пауза между итерациями воркера")
    
    ap.add_argument("--global-interval", type=float, default=0.08, help="Глобальный интервал между запросами (сек)")
    ap.add_argument("--verbose", action="store_true", help="Подробные логи")
    
    args = ap.parse_args()

    # Загрузка прокси
    try:
        proxies = load_proxies(args.proxies)
    except Exception as e:
        print(f"Ошибка загрузки прокси: {e}")
        return

    if not proxies:
        print("Список прокси пуст.")
        return

    # Загрузка имен
    names = load_names(args.names)
    if not names:
        print(f"Список имен пуст или файл {args.names} не найден.")
        return
    
    # Resume логика
    processed_names = set()
    if os.path.exists(args.out):
        print(f"Файл {args.out} найден, читаем уже обработанные предметы...")
        processed_names = load_processed_names(args.out)
        print(f"Найдено {len(processed_names)} обработанных предметов.")
        
    names_to_process = [n for n in names if n not in processed_names]
    
    if not names_to_process:
        print("Все предметы из names.txt уже есть в файле. Нечего делать.")
        return

    print(f"Загружено всего имен: {len(names)}. Осталось обработать: {len(names_to_process)}")
    print(f"Прокси: {len(proxies)} шт.")
    print(f"Потоки: {args.workers}")
    print(f"Валюта ID: {args.currency} (1=USD)")

    # Очередь задач
    tasks_q: "queue.Queue[Task]" = queue.Queue()
    for n in names_to_process:
        tasks_q.put(Task(name=n, tries_left=5))

    # Сделаем больше проходов по списку прокси (например 10), чтобы при большом кол-ве предметов прокси не кончались слишком быстро
    # Если прокси забанят (429/403), он все равно выбывает до конца скрипта (точнее меняется, но если все забанены - работа встанет сама)
    proxy_pool = ProxyPool(proxies, max_passes=20) 
# 1 проход по прокси достаточно, если их хватает
    # Если прокси мало, а предметов много, можно увеличить max_passes, или просто 1
    # В данном ТЗ не уточнялось, оставим 1

    results_q: "queue.Queue[Optional[PriceRow]]" = queue.Queue()
    rate_limiter = GlobalRateLimiter(args.global_interval)

    stats = {"processed": 0, "ok": 0, "fail": 0, "429": 0, "403": 0, "other_err": 0}
    stats_lock = threading.Lock()
    stop_evt = threading.Event()

    # Writer
    writer_t = threading.Thread(target=excel_writer, args=(args.out, args.append, results_q), daemon=True)
    writer_t.start()
    
    # Progress
    prog_t = threading.Thread(target=progress_printer, args=(5.0, tasks_q, stats, stats_lock, stop_evt), daemon=True)
    prog_t.start()

    # Workers
    threads = []
    for i in range(args.workers):
        t = threading.Thread(
            target=worker,
            args=(
                i,
                tasks_q,
                results_q,
                proxy_pool,
                args.appid,
                args.currency,
                args.proxy_scheme,
                args.max_req_per_proxy,
                args.timeout,
                args.delay,
                rate_limiter,
                stats,
                stats_lock,
                args.verbose,
                stop_evt,
            ),
            daemon=True,
        )
        t.start()
        threads.append(t)

    # Wait for completion
    while True:
        if tasks_q.unfinished_tasks == 0:
            break
        if stop_evt.is_set():
             # drain
            while not tasks_q.empty():
                try:
                    tasks_q.get_nowait()
                    tasks_q.task_done()
                except:
                    break
            break
        time.sleep(1.0)

    stop_evt.set()
    for t in threads:
        t.join(timeout=2.0)

    results_q.put(None)
    writer_t.join()

    print("Готово!")
    print(stats)


if __name__ == "__main__":
    main()
