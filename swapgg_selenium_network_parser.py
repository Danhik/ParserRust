#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
swapgg_selenium_network_parser.py

Selenium (browser) скроллит swap.gg, а данные берутся из Network (XHR JSON) /api/inventory/site.

Идея:
  - Открываем https://swap.gg/ru/trade
  - Переключаем игру на Rust (важно: кнопка на стороне площадки/market)
  - Находим правый vue-recycle-scroller (инвентарь площадки)
  - Скроллим, чтобы сайт сам делал XHR-запросы
  - Перехватываем ответы XHR через selenium-wire
  - Парсим JSON и пишем в Excel

Запуск:
  python swapgg_selenium_network_parser.py --out swap_rust.xlsx --max-items 2000 --verbose

Важно:
  - Headless лучше выключить (по умолчанию выключен).
  - Можно указать --user-data-dir ./chrome_profile чтобы сохранялась сессия.
"""

import argparse
import json
import os
import time
from dataclasses import dataclass
from datetime import datetime, timezone
from typing import Dict, List, Optional, Set, Tuple

from openpyxl import Workbook, load_workbook

from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# selenium-wire
from seleniumwire import webdriver


# -----------------------------
# Models
# -----------------------------

@dataclass(frozen=True)
class ItemRow:
    fetched_at_utc: str
    skip: int
    name: str
    price_usd: str
    product_id: str
    stack: str


# -----------------------------
# Excel helpers
# -----------------------------

def ensure_workbook(out_path: str, append: bool):
    if append and os.path.exists(out_path):
        wb = load_workbook(out_path)
        ws = wb.active
        if ws.max_row < 1:
            ws.append(["fetched_at_utc", "skip", "name", "price_usd", "product_id", "stack"])
        return wb, ws

    wb = Workbook()
    ws = wb.active
    ws.title = "items"
    ws.append(["fetched_at_utc", "skip", "name", "price_usd", "product_id", "stack"])
    return wb, ws


def load_existing_keys_from_xlsx(path: str) -> Set[str]:
    """
    Для append: чтобы не писать дубликаты.
    key = product_id|stack|price_usd (достаточно стабильно)
    """
    if not os.path.exists(path):
        return set()

    wb = load_workbook(path)
    ws = wb.active
    seen = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 6:
            continue
        product_id = row[4] or ""
        stack = row[5] or ""
        price = row[3] or ""
        if product_id and stack:
            seen.add(f"{product_id}|{stack}|{price}")
    return seen


def append_rows_to_xlsx(out_path: str, rows: List[ItemRow], append: bool, seen_keys: Set[str], save_every: int = 200, verbose: bool = False):
    wb, ws = ensure_workbook(out_path, append)
    written = 0
    for r in rows:
        key = f"{r.product_id}|{r.stack}|{r.price_usd}"
        if key in seen_keys:
            continue
        ws.append([r.fetched_at_utc, r.skip, r.name, r.price_usd, r.product_id, r.stack])
        seen_keys.add(key)
        written += 1
        if written % save_every == 0:
            wb.save(out_path)
            if verbose:
                print(f"[xlsx] saved +{written} rows -> {out_path}")
    wb.save(out_path)
    if verbose:
        print(f"[xlsx] final save, appended={written} -> {out_path}")


# -----------------------------
# Page helpers
# -----------------------------

def looks_like_cloudflare_challenge_title(driver) -> bool:
    t = (driver.title or "").lower()
    return "just a moment" in t or "checking your browser" in t


def wait_until_ready(driver, timeout_s: int, verbose: bool = False):
    """
    Ждём пока появятся элементы интерфейса.
    Если CF — просто ждём (в обычном видимом окне можно руками пройти).
    """
    t0 = time.time()
    while time.time() - t0 < timeout_s:
        if looks_like_cloudflare_challenge_title(driver):
            if verbose:
                print("[info] Похоже на CF (Just a moment...). Если в окне есть проверка — пройди её вручную.")
            time.sleep(1.5)
            continue
        try:
            WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "button[title], .vue-recycle-scroller"))
            )
            return
        except Exception:
            time.sleep(0.8)
    if verbose:
        print("[warn] timeout waiting page ready")


def accept_cookies_if_present(driver, verbose: bool = False):
    # очень мягкая эвристика
    xps = [
        "//button[contains(translate(., 'ACEGPT', 'acegpt'), 'accept')]",
        "//button[contains(translate(., 'ACEGPT', 'acegpt'), 'agree')]",
        "//button[contains(., 'Принять')]",
        "//button[contains(., 'Соглас')]",
        "//button[contains(., 'OK')]",
        "//button[contains(., 'Ок')]",
    ]
    for xp in xps:
        try:
            btn = WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.XPATH, xp)))
            driver.execute_script("arguments[0].click();", btn)
            if verbose:
                print("[info] cookies accepted")
            time.sleep(0.3)
            return
        except Exception:
            pass


def _safe_click(driver, el, verbose=False) -> bool:
    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center', inline:'center'});", el)
        time.sleep(0.1)
        driver.execute_script("arguments[0].click();", el)
        return True
    except Exception:
        try:
            el.click()
            return True
        except Exception as e:
            if verbose:
                print(f"[warn] click failed: {e}")
            return False


def select_game_rust_market_side(driver, wait_s: int = 20, verbose: bool = False) -> bool:
    """
    На странице 2 кнопки Rust: слева (инвентарь пользователя) и справа (инвентарь площадки).
    Нам нужно нажать ПРАВУЮ (максимальный X).
    """
    try:
        WebDriverWait(driver, wait_s).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'button[title]')))
    except Exception:
        if verbose:
            print("[warn] buttons not found")
        return False

    buttons = driver.find_elements(By.CSS_SELECTOR, 'button[title="Rust"]')
    if not buttons:
        # fallback по тексту
        buttons = driver.find_elements(By.XPATH, "//button[contains(., 'Rust')]")

    if not buttons:
        if verbose:
            print("[warn] Rust button not found")
        return False

    # выбираем "самую правую" по X
    best = None
    best_x = -1e9
    for b in buttons:
        try:
            x = driver.execute_script("return arguments[0].getBoundingClientRect().x;", b)
            if x is not None and x > best_x:
                best_x = x
                best = b
        except Exception:
            continue

    if best is None:
        best = buttons[-1]

    ok = _safe_click(driver, best, verbose=verbose)
    if verbose:
        print(f"[info] clicked Rust on market side: ok={ok} x={best_x:.1f}")

    # иногда полезно нажать обе — но безопасно только если хочешь
    # (если хочешь — раскомментируй)
    # for b in buttons:
    #     _safe_click(driver, b, verbose=False)
    #     time.sleep(0.15)

    time.sleep(1.0)
    return ok


def find_right_scroller(driver, wait_s: int = 20, verbose: bool = False):
    """
    На странице бывает несколько .vue-recycle-scroller
    Берём самый правый по X — это обычно инвентарь площадки.
    """
    try:
        WebDriverWait(driver, wait_s).until(EC.presence_of_element_located((By.CSS_SELECTOR, ".vue-recycle-scroller")))
    except Exception:
        if verbose:
            print("[warn] scroller not found")
        return None

    scrollers = driver.find_elements(By.CSS_SELECTOR, ".vue-recycle-scroller")
    if not scrollers:
        return None

    best = None
    best_x = -1e9
    for s in scrollers:
        try:
            x = driver.execute_script("return arguments[0].getBoundingClientRect().x;", s)
            w = driver.execute_script("return arguments[0].getBoundingClientRect().width;", s)
            if verbose:
                print(f"[debug] scroller x={x:.1f} w={w:.1f}")
            if x is not None and x > best_x:
                best_x = x
                best = s
        except Exception:
            continue

    if best is None:
        best = scrollers[-1]

    if verbose:
        print(f"[info] selected right scroller x={best_x:.1f}")
    return best


# -----------------------------
# Network parsing via selenium-wire
# -----------------------------

def _decode_body(resp_body: bytes) -> str:
    # selenium-wire обычно отдаёт уже распакованное, но подстрахуемся
    try:
        return resp_body.decode("utf-8", errors="replace")
    except Exception:
        return str(resp_body)


def _parse_request_skip(req) -> int:
    """
    В POST body лежит JSON:
      {"data":{"force":false},"metadata":{"query":{...,"pagination":{"limit":50,"skip":0}}}}
    """
    try:
        if not req.body:
            return -1
        raw = req.body
        if isinstance(raw, (bytes, bytearray)):
            raw = raw.decode("utf-8", errors="replace")
        data = json.loads(raw)
        pag = (((data.get("metadata") or {}).get("query") or {}).get("pagination") or {})
        sk = pag.get("skip")
        return int(sk) if isinstance(sk, int) else -1
    except Exception:
        return -1


def extract_items_from_inventory_response(json_obj: dict, skip: int) -> List[ItemRow]:
    out: List[ItemRow] = []
    fetched_at = datetime.now(timezone.utc).isoformat()

    data = json_obj.get("data") or {}
    items = data.get("items") or []
    if not isinstance(items, list):
        return out

    for it in items:
        try:
            product = it.get("product") or {}
            name = (product.get("name") or "").strip()
            product_id = (product.get("_id") or "").strip()
            stack = (it.get("stack") or "").strip()
            price_int = it.get("price")

            # price: 75417 => $754.17
            price_usd = ""
            if isinstance(price_int, int):
                price_usd = f"${price_int / 100.0:.2f}"

            if name and product_id and stack:
                out.append(ItemRow(
                    fetched_at_utc=fetched_at,
                    skip=skip,
                    name=name,
                    price_usd=price_usd,
                    product_id=product_id,
                    stack=stack
                ))
        except Exception:
            continue

    return out


# -----------------------------
# Scrolling loop + network harvesting
# -----------------------------

def harvest_by_scrolling_network(
    driver,
    scroller_el,
    max_items: int,
    scroll_pause: float,
    stable_rounds: int,
    verbose: bool = False
) -> List[ItemRow]:
    """
    Скроллим scroller. После каждого шага читаем новые network requests из selenium-wire.
    Остановка:
      - набрали max_items (если >0)
      - stable_rounds раз подряд не добавили ни одного нового item из network
    """
    collected: List[ItemRow] = []
    seen: Set[str] = set()  # product_id|stack|price
    last_req_index = 0
    stable = 0

    def pull_new_from_requests() -> int:
        nonlocal last_req_index
        added = 0

        reqs = driver.requests  # selenium-wire list
        if last_req_index >= len(reqs):
            return 0

        new_part = reqs[last_req_index:]
        last_req_index = len(reqs)

        for r in new_part:
            try:
                if not r.response:
                    continue
                if "/api/inventory/site" not in (r.url or ""):
                    continue

                status = getattr(r.response, "status_code", None)
                if status not in (200, 201):
                    continue

                # content-type должен быть json
                ct = (r.response.headers.get("Content-Type") or "").lower()
                if "application/json" not in ct:
                    # иногда CF может вернуть html — отфильтруем
                    body_txt = _decode_body(r.response.body or b"")[:200].lower()
                    if "<html" in body_txt or "just a moment" in body_txt:
                        continue

                skip = _parse_request_skip(r)
                body_txt = _decode_body(r.response.body or b"")
                js = json.loads(body_txt)

                items = extract_items_from_inventory_response(js, skip=skip)
                for it in items:
                    key = f"{it.product_id}|{it.stack}|{it.price_usd}"
                    if key in seen:
                        continue
                    seen.add(key)
                    collected.append(it)
                    added += 1

            except Exception:
                continue

        return added

    # начальная выборка (если запрос уже прилетел)
    added0 = pull_new_from_requests()
    if verbose:
        print(f"[info] start: collected={len(collected)} (+{added0} from network)")

    while True:
        if max_items > 0 and len(collected) >= max_items:
            if verbose:
                print(f"[info] reached max-items={max_items}")
            break

        # Скроллим быстрее: прыжок почти на высоту контейнера
        try:
            driver.execute_script(
                "arguments[0].scrollTop = arguments[0].scrollTop + Math.floor(arguments[0].clientHeight * 0.95);",
                scroller_el
            )
        except Exception:
            try:
                driver.execute_script("arguments[0].scrollTop = arguments[0].scrollHeight;", scroller_el)
            except Exception:
                pass

        time.sleep(scroll_pause)

        added = pull_new_from_requests()
        if verbose:
            print(f"[info] collected={len(collected)} (+{added}) stable={stable}/{stable_rounds}")

        if added == 0:
            stable += 1
        else:
            stable = 0

        if stable >= stable_rounds:
            if verbose:
                print("[info] stable reached (no new network items). stop.")
            break

        # чтобы память не росла бесконечно:
        # чистим requests иногда (тогда нужно сбросить last_req_index)
        if len(driver.requests) > 2000:
            try:
                driver.requests.clear()
                last_req_index = 0
                if verbose:
                    print("[debug] cleared driver.requests to save memory")
            except Exception:
                pass

    return collected


# -----------------------------
# Driver setup
# -----------------------------

def build_driver(headless: bool, user_data_dir: Optional[str], verbose: bool):
    opts = ChromeOptions()
    if headless:
        opts.add_argument("--headless=new")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--window-size=1400,900")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--lang=ru-RU")
    opts.add_argument("--disable-blink-features=AutomationControlled")

    if user_data_dir:
        opts.add_argument(f"--user-data-dir={user_data_dir}")
        if verbose:
            print(f"[info] using user-data-dir: {user_data_dir}")

    # selenium-wire options (можно не трогать)
    sw_opts = {
        "disable_encoding": True,  # чтобы тело ответа было проще читать
    }

    return webdriver.Chrome(options=opts, seleniumwire_options=sw_opts)


# -----------------------------
# Main
# -----------------------------

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--url", default="https://swap.gg/ru/trade", help="Страница trade")
    ap.add_argument("--out", default="swap_network.xlsx", help="Выходной Excel файл")
    ap.add_argument("--append", action="store_true", help="Дописать в существующий Excel")

    ap.add_argument("--max-items", type=int, default=2000, help="Лимит предметов (0=без лимита, осторожно)")
    ap.add_argument("--scroll-pause", type=float, default=0.25, help="Пауза между скроллами (чем меньше — тем быстрее)")
    ap.add_argument("--stable-rounds", type=int, default=30, help="Стоп если N шагов подряд нет новых из network")

    ap.add_argument("--challenge-wait", type=int, default=120, help="Сколько секунд ждать ручной CF-проверки")
    ap.add_argument("--headless", action="store_true", help="Headless режим (лучше НЕ включать)")
    ap.add_argument("--user-data-dir", default="", help="Путь к профилю Chrome, например ./chrome_profile")
    ap.add_argument("--verbose", action="store_true", help="Подробные логи")

    args = ap.parse_args()
    user_data_dir = args.user_data_dir.strip() or None

    # append keys
    seen_keys = load_existing_keys_from_xlsx(args.out) if args.append else set()

    driver = build_driver(headless=args.headless, user_data_dir=user_data_dir, verbose=args.verbose)

    try:
        if args.verbose:
            print(f"[info] open: {args.url}")

        driver.get(args.url)
        wait_until_ready(driver, timeout_s=args.challenge_wait, verbose=args.verbose)
        accept_cookies_if_present(driver, verbose=args.verbose)

        # На всякий случай: если была "Just a moment..."
        if looks_like_cloudflare_challenge_title(driver):
            if args.verbose:
                print("[info] still CF page, waiting again...")
            wait_until_ready(driver, timeout_s=args.challenge_wait, verbose=args.verbose)

        # Переключаем market side на Rust
        select_game_rust_market_side(driver, wait_s=25, verbose=args.verbose)

        # Находим правый scroller
        scroller = find_right_scroller(driver, wait_s=25, verbose=args.verbose)
        if scroller is None:
            print("[error] Не найден vue-recycle-scroller (правый).")
            return

        # Дадим странице чуть времени, чтобы первый XHR успел прилететь
        time.sleep(1.0)

        # Сбор через network во время скролла
        rows = harvest_by_scrolling_network(
            driver=driver,
            scroller_el=scroller,
            max_items=args.max_items,
            scroll_pause=args.scroll_pause,
            stable_rounds=args.stable_rounds,
            verbose=args.verbose
        )

        if args.verbose:
            print(f"[info] total rows collected from network: {len(rows)}")

        # запись
        append_rows_to_xlsx(args.out, rows, append=args.append, seen_keys=seen_keys, save_every=200, verbose=args.verbose)

        print(f"Готово. Файл: {args.out}")

    finally:
        try:
            driver.quit()
        except Exception:
            pass


if __name__ == "__main__":
    main()
