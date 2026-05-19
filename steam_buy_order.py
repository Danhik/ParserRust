import os
import sys
import time
import json
import base64
import argparse
import subprocess
from typing import List, Optional, Tuple, Dict, Any
from dataclasses import dataclass
from datetime import datetime

import requests
from openpyxl import load_workbook

# =============================================================
#  НАСТРОЙКИ (без авторизации — она в профилях)
# =============================================================

DEFAULT_APPID         = 252490   # Rust (730 = CS2)
DEFAULT_CURRENCY      = 37       # 37=KZT | 1=USD | 5=RUB | 3=EUR

DEFAULT_EXCEL_FILE    = "orders.xlsx"
DEFAULT_PROXIES_FILE  = "proxies.txt"
USE_PROXIES           = False

# Тайминги
REQUEST_TIMEOUT       = 20       # сек на запрос
DELAY_BETWEEN_ORDERS  = 1.5      # сек между ордерами
DELAY_ON_RATELIMIT    = 15.0     # сек при 429
MAX_RETRIES           = 3        # попыток на ордер
MAX_ORDERS_PER_RUN    = 0        # 0 = все

# Профили
SESSIONS_DIR          = "steam_sessions"   # папка с профилями

# =============================================================
#  КОНСТАНТЫ
# =============================================================

URL_BUY_ORDER    = "https://steamcommunity.com/market/createbuyorder/"
URL_MY_LISTINGS  = "https://steamcommunity.com/market/mylistings/?norender=1"
STEAM_LOGIN_URL  = "https://steamcommunity.com/login/home/"

CURRENCY_NAMES = {
    1: "USD", 2: "GBP", 3: "EUR", 4: "CHF", 5: "RUB",
    6: "PLN", 8: "JPY", 17: "TRY", 18: "UAH", 37: "KZT",
}

STEAM_ERESULT = {
    1:   (True,  "OK"),
    2:   (False, "Общая ошибка Steam"),
    8:   (False, "Неверный параметр (цена или валюта не совпадает с аккаунтом)"),
    9:   (False, "Предмет не найден на маркете"),
    16:  (False, "Таймаут Steam"),
    25:  (False, "Превышен лимит buy-ордеров на аккаунте"),
    29:  (False, "Дубликат: ордер на этот предмет уже существует"),
    42:  (False, "Предмет не найден (проверь market_hash_name)"),
    84:  (False, "Rate limit: слишком много покупок без подтверждения"),
    107: (False, "Недостаточно средств на Steam кошельке"),
}
# При этих кодах повтор бесполезен — сразу к следующему ордеру
NO_RETRY_CODES = {9, 25, 29, 42, 107}
# При этих кодах нужна пауза перед следующим ордером
RATELIMIT_ERESULT_CODES = {84}

# =============================================================
#  DATACLASS
# =============================================================

@dataclass
class OrderTask:
    market_hash_name: str
    price_cents: int   # в единицах валюты * 100
    quantity: int = 1
    appid: int = DEFAULT_APPID


# =============================================================
#  UTILS
# =============================================================

def ts() -> str:
    return datetime.now().strftime("%H:%M:%S")


def fmt_price(cents: int, currency: int = DEFAULT_CURRENCY) -> str:
    return f"{cents / 100:.2f} {CURRENCY_NAMES.get(currency, '?')}"


def load_proxies(filepath: str) -> List[str]:
    if not os.path.exists(filepath):
        return []
    lines = [l.strip() for l in open(filepath, encoding="utf-8") if l.strip()]
    print(f"[+] Прокси: {len(lines)}")
    return lines


def format_proxy(proxy_str: str, scheme: str = "http") -> dict:
    if not proxy_str:
        return {}
    parts = proxy_str.split(":")
    if len(parts) == 4:
        ip, port, user, pwd = parts
        url = f"{scheme}://{user}:{pwd}@{ip}:{port}"
    elif len(parts) == 2:
        url = f"{scheme}://{proxy_str}"
    else:
        url = f"{scheme}://{proxy_str}"
    return {"http": url, "https": url}


# =============================================================
#  EXCEL READER
# =============================================================

def load_orders_from_excel(filepath: str) -> List[OrderTask]:
    """
    Колонка A: market_hash_name
    Колонка B: цена (число, в валюте аккаунта)
    Колонка C: количество (опц., по умолч. 1)
    Колонка D: appid (опц.)
    Строка 1 — заголовок, пропускается автоматически.
    """
    if not os.path.exists(filepath):
        print(f"[!] Файл '{filepath}' не найден.")
        return []

    wb = load_workbook(filepath, data_only=True)
    ws = wb.active
    orders = []

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if not row or row[0] is None:
            continue
        name = str(row[0]).strip()
        if not name or name.lower() in ("name", "item name", "название", "предмет"):
            continue
        try:
            price_float = float(row[1])
        except (TypeError, ValueError, IndexError):
            print(f"  [!] Строка {row_idx}: некорректная цена, пропускаем.")
            continue

        qty = 1
        if len(row) > 2 and row[2] is not None:
            try:
                qty = max(1, int(row[2]))
            except (TypeError, ValueError):
                pass

        appid = DEFAULT_APPID
        if len(row) > 3 and row[3] is not None:
            try:
                appid = int(row[3])
            except (TypeError, ValueError):
                pass

        orders.append(OrderTask(
            market_hash_name=name,
            price_cents=int(round(price_float * 100)),
            quantity=qty,
            appid=appid
        ))

    print(f"[+] Загружено ордеров: {len(orders)}")
    return orders


# =============================================================
#  JWT DECODE (без сети)
# =============================================================

def decode_jwt_exp(steam_login_secure: str) -> Optional[int]:
    """Возвращает Unix timestamp истечения JWT или None."""
    try:
        jwt_part = steam_login_secure.split("||")[-1] if "||" in steam_login_secure else steam_login_secure
        parts = jwt_part.split(".")
        if len(parts) < 2:
            return None
        payload_b64 = parts[1]
        padding = 4 - len(payload_b64) % 4
        if padding != 4:
            payload_b64 += "=" * padding
        payload = json.loads(base64.urlsafe_b64decode(payload_b64))
        return payload.get("exp")
    except Exception:
        return None


def check_token_expiry(steam_login_secure: str) -> Tuple[bool, Optional[int]]:
    exp = decode_jwt_exp(steam_login_secure)
    if exp is None:
        return False, None
    secs = exp - int(time.time())
    return secs > 0, secs


# =============================================================
#  ПРОФИЛИ (сохранение/загрузка куков)
# =============================================================

def _session_path(profile: str) -> str:
    os.makedirs(SESSIONS_DIR, exist_ok=True)
    return os.path.join(SESSIONS_DIR, f"steam_session_{profile}.json")


def load_profile(profile: str) -> Dict[str, Any]:
    path = _session_path(profile)
    if not os.path.exists(path):
        return {}
    try:
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def save_profile(profile: str, data: Dict[str, Any]) -> None:
    path = _session_path(profile)
    data["saved_at"] = time.time()
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    print(f"[+] Профиль '{profile}' сохранён -> {path}")


def list_profiles() -> List[str]:
    if not os.path.isdir(SESSIONS_DIR):
        return []
    return [
        f.replace("steam_session_", "").replace(".json", "")
        for f in os.listdir(SESSIONS_DIR)
        if f.startswith("steam_session_") and f.endswith(".json")
    ]


# =============================================================
#  ПРОВЕРКА СЕССИИ (сетевая)
# =============================================================

def validate_session_online(session_id: str, steam_login_secure: str,
                             currency: int, timeout: int = 10) -> bool:
    sess = _build_session(session_id, steam_login_secure, currency)
    try:
        r = sess.get(URL_MY_LISTINGS, timeout=timeout, allow_redirects=False)
        if r.status_code == 200:
            return r.json().get("success") is True
        return False
    except Exception:
        return False


# =============================================================
#  БРАУЗЕРНАЯ АВТОРИЗАЦИЯ
# =============================================================

def _get_chrome_major() -> Optional[int]:
    paths = [
        r"C:\Program Files\Google\Chrome\Application\chrome.exe",
        r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
    ]
    for p in paths:
        if os.path.exists(p):
            try:
                out = subprocess.check_output(
                    ["powershell", "-Command", f'(Get-Item "{p}").VersionInfo.ProductVersion'],
                    stderr=subprocess.DEVNULL, timeout=5
                ).decode().strip()
                return int(out.split(".")[0])
            except Exception:
                pass
    return None


def auth_via_browser(profile: str) -> Dict[str, Any]:
    """
    Открывает Chrome на странице входа Steam.
    Ждёт пока пользователь войдёт и нажмёт Enter в консоли.
    Извлекает sessionid и steamLoginSecure, сохраняет профиль.
    """
    try:
        import undetected_chromedriver as uc
    except ImportError:
        print("[!] undetected-chromedriver не установлен.")
        print("    Установи: pip install undetected-chromedriver")
        return {}

    print(f"\n[Browser] Запуск Chrome для профиля '{profile}'...")
    options = uc.ChromeOptions()
    major = _get_chrome_major()
    try:
        driver = uc.Chrome(options=options, version_main=major) if major else uc.Chrome(options=options)
    except Exception as e:
        print(f"[Browser] Ошибка запуска Chrome: {e}")
        return {}

    try:
        driver.get(STEAM_LOGIN_URL)
        print("[Browser] Страница входа открыта.")
        print("[Browser] Войдите в Steam аккаунт в открывшемся окне браузера.")
        print()

        input("         Нажмите ENTER после успешного входа в аккаунт...")
        print("[Browser] Считываем куки...")

        # Собираем ВСЕ куки steamcommunity.com
        all_cookies = {c["name"]: c["value"] for c in driver.get_cookies()}
        ua = driver.execute_script("return navigator.userAgent;")

        session_id         = all_cookies.get("sessionid", "")
        steam_login_secure = all_cookies.get("steamLoginSecure", "")

        if not session_id or not steam_login_secure:
            print("[Browser] Куки не найдены. Убедитесь что вход выполнен успешно.")
            return {}

        data = {
            "session_id":          session_id,
            "steam_login_secure":  steam_login_secure,
            "user_agent":          ua,
            "all_cookies":         all_cookies,
        }
        save_profile(profile, data)
        print("[Browser] Авторизация успешна!")
        return data

    finally:
        try:
            driver.quit()
        except Exception:
            pass  # Подавляем OSError WinError 6 при закрытии


# =============================================================
#  ПОЛУЧЕНИЕ / ОБНОВЛЕНИЕ СЕССИИ
# =============================================================

def ensure_session(profile: str, force_reauth: bool = False,
                   skip_validate: bool = False) -> Tuple[str, str, int, Optional[Dict], Optional[str]]:
    """
    Возвращает (session_id, steam_login_secure, currency, all_cookies, user_agent).
    all_cookies/user_agent — из браузера, нужны для передачи всех куков в _build_session.
    Если куки невалидны — запускает браузерную авторизацию.
    """
    data = {} if force_reauth else load_profile(profile)

    if data:
        session_id         = data.get("session_id", "")
        steam_login_secure = data.get("steam_login_secure", "")
        currency           = data.get("currency", DEFAULT_CURRENCY)
        all_cookies        = data.get("all_cookies") or None
        user_agent         = data.get("user_agent") or None

        print(f"\n[*] Профиль '{profile}' — проверка токена...")
        is_valid, secs = check_token_expiry(steam_login_secure)

        if not is_valid:
            print("  [ERR] JWT истёк — нужна переавторизация.")
            data = {}
        else:
            days = secs // 86400
            hrs  = (secs % 86400) // 3600
            print(f"  [OK] JWT действителен ещё ~{days}д {hrs}ч")

            if not skip_validate:
                print("[*] Проверка сессии через Steam API...")
                ok = validate_session_online(session_id, steam_login_secure, currency)
                if ok:
                    print("  [OK] Сессия активна.")
                    return session_id, steam_login_secure, currency, all_cookies, user_agent
                else:
                    print("  [ERR] Сессия не принята Steam (sessionid протух).")
                    data = {}
            else:
                return session_id, steam_login_secure, currency, all_cookies, user_agent
    else:
        if not force_reauth:
            print(f"\n[*] Профиль '{profile}' не найден.")

    # Нужна браузерная авторизация
    print(f"\n[*] Запускаем авторизацию для профиля '{profile}'...")
    data = auth_via_browser(profile)
    if not data:
        print("[!] Авторизация не удалась. Выход.")
        sys.exit(1)

    session_id         = data["session_id"]
    steam_login_secure = data["steam_login_secure"]
    all_cookies        = data.get("all_cookies") or {}
    user_agent         = data.get("user_agent") or None

    currency_raw = all_cookies.get("steamCurrencyId", str(DEFAULT_CURRENCY))
    try:
        currency = int(currency_raw)
    except ValueError:
        currency = DEFAULT_CURRENCY

    print(f"  [+] Валюта аккаунта: {CURRENCY_NAMES.get(currency, currency)} ({currency})")

    data["currency"] = currency
    save_profile(profile, data)

    return session_id, steam_login_secure, currency, all_cookies, user_agent


# =============================================================
#  STEAM SESSION
# =============================================================

def _build_session(session_id: str, steam_login_secure: str,
                   currency: int,
                   all_cookies: Optional[Dict[str, str]] = None,
                   user_agent: Optional[str] = None) -> requests.Session:
    """
    Строит requests.Session.
    Если переданы all_cookies (из браузерной авторизации) — применяет ВСЕ куки
    включая webTradeEligibility, steamCountry и т.д. (нужно для HTTP 406).
    """
    ua = user_agent or (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/148.0.0.0 Safari/537.36"
    )
    sess = requests.Session()
    sess.headers.update({
        "User-Agent":         ua,
        "Accept":             "*/*",
        "Accept-Language":    "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7",
        "Accept-Encoding":    "gzip, deflate, br",
        "Origin":             "https://steamcommunity.com",
        "Sec-Fetch-Dest":     "empty",
        "Sec-Fetch-Mode":     "cors",
        "Sec-Fetch-Site":     "same-origin",
        "sec-ch-ua":          '"Chromium";v="148", "Google Chrome";v="148", "Not/A)Brand";v="99"',
        "sec-ch-ua-mobile":   "?0",
        "sec-ch-ua-platform": '"Windows"',
    })
    d = "steamcommunity.com"

    if all_cookies:
        # Применяем все браузерные куки (webTradeEligibility, steamCountry и др.)
        for name, value in all_cookies.items():
            sess.cookies.set(name, value, domain=d)
    else:
        # Минимальный набор (если профиль создан вручную)
        sess.cookies.set("sessionid",        session_id,         domain=d)
        sess.cookies.set("steamLoginSecure", steam_login_secure, domain=d)
        sess.cookies.set("Steam_Language",   "english",          domain=d)
        sess.cookies.set("steamCurrencyId",  str(currency),      domain=d)
        sess.cookies.set("timezoneOffset",   "10800,0",          domain=d)
        sess.cookies.set("timezoneName",     "Europe/Moscow",    domain=d)

    return sess


# =============================================================
#  ВЫСТАВЛЕНИЕ ОРДЕРА
# =============================================================

def place_buy_order(sess: requests.Session, order: OrderTask,
                    session_id: str, currency: int,
                    proxy_dict: dict, timeout: int) -> dict:
    referer = (
        f"https://steamcommunity.com/market/listings/"
        f"{order.appid}/{requests.utils.quote(order.market_hash_name)}?buy=1"
    )
    form_data = {
        "sessionid":           (None, session_id),
        "currency":            (None, str(currency)),
        "appid":               (None, str(order.appid)),
        "market_hash_name":    (None, order.market_hash_name),
        "price_total":         (None, str(order.price_cents)),
        "tradefee_tax":        (None, "0"),
        "quantity":            (None, str(order.quantity)),
        "first_name":          (None, ""),
        "last_name":           (None, ""),
        "billing_address":     (None, ""),
        "billing_address_two": (None, ""),
        "billing_country":     (None, "KZ"),
        "billing_city":        (None, ""),
        "billing_state":       (None, ""),
        "billing_postal_code": (None, ""),
        "confirmation":        (None, "0"),
        "save_my_address":     (None, "1"),
    }
    _fail = lambda msg, sr=0, nr=False: {
        "ok": False, "order_id": None, "code": None,
        "steam_result": sr, "no_retry": nr, "error": msg
    }
    try:
        r = sess.post(
            URL_BUY_ORDER,
            files=form_data,
            headers={"Referer": referer, "x-valve-request-type": "buyAction"},
            proxies=proxy_dict,
            timeout=timeout,
        )
        if r.status_code == 200:
            data = r.json()
            sr = data.get("success", 0)
            if sr == 1:
                return {"ok": True, "order_id": data.get("buy_orderid"),
                        "code": 200, "steam_result": 1, "no_retry": False, "error": None}
            _, desc = STEAM_ERESULT.get(sr, (False, f"EResult={sr}"))
            msg = data.get("message") or desc
            return {"ok": False, "order_id": None, "code": 200,
                    "steam_result": sr, "no_retry": sr in NO_RETRY_CODES,
                    "error": f"EResult {sr}: {desc} | {msg}"}
        return _fail(f"HTTP {r.status_code}")
    except requests.exceptions.Timeout:
        return _fail("Timeout")
    except requests.exceptions.ProxyError as e:
        return _fail(f"ProxyError: {e}")
    except Exception as e:
        return _fail(str(e))


# =============================================================
#  ГЛАВНАЯ ЛОГИКА
# =============================================================

def run(orders: List[OrderTask], session_id: str, steam_login_secure: str,
        currency: int, proxies: List[str], proxy_scheme: str,
        delay: float, retries: int, timeout: int, max_orders: int,
        verbose: bool, profile: str,
        all_cookies: Optional[Dict[str, str]] = None,
        user_agent: Optional[str] = None):

    sess = _build_session(session_id, steam_login_secure, currency,
                          all_cookies=all_cookies, user_agent=user_agent)
    proxy_idx = 0
    results = []
    total = len(orders) if not max_orders else min(len(orders), max_orders)
    cur_name = CURRENCY_NAMES.get(currency, str(currency))

    print(f"\n{'='*58}")
    print(f"  Steam Buy Order Placer")
    print(f"  Профиль: {profile}  |  AppID: {DEFAULT_APPID}  |  Валюта: {cur_name}")
    print(f"  Ордеров: {total}  |  Задержка: {delay}с  |  Попыток: {retries}")
    print(f"{'='*58}\n")

    for i, order in enumerate(orders[:total], 1):
        price_str = fmt_price(order.price_cents, currency)
        print(f"[{ts()}] [{i}/{total}] '{order.market_hash_name}' x{order.quantity} @ {price_str}")

        proxy_str = ""
        if proxies and USE_PROXIES:
            proxy_str = proxies[proxy_idx % len(proxies)]
            proxy_idx += 1
        proxy_dict = format_proxy(proxy_str, proxy_scheme)

        if verbose and proxy_str:
            print(f"  -> Прокси: {proxy_str.split(':')[0]}")

        placed = False
        for attempt in range(1, retries + 1):
            res = place_buy_order(sess, order, session_id, currency, proxy_dict, timeout)

            if res["ok"]:
                print(f"  [OK] Выставлен! order_id = {res['order_id']}")
                results.append({**vars(order), "order_id": res["order_id"], "status": "OK"})
                placed = True
                break

            code = res.get("code")
            err  = res.get("error", "?")

            if code == 403:
                print("  [ERR] [403] Куки недействительны. Запусти с --reauth.")
                results.append({**vars(order), "order_id": None, "status": "AUTH_ERROR"})
                _print_summary(results, currency)
                return results

            if code == 429:
                print(f"  [!!] Rate limit. Пауза {DELAY_ON_RATELIMIT}с... ({attempt}/{retries})")
                time.sleep(DELAY_ON_RATELIMIT)
                if proxies and USE_PROXIES:
                    proxy_str = proxies[proxy_idx % len(proxies)]
                    proxy_idx += 1
                    proxy_dict = format_proxy(proxy_str, proxy_scheme)
                continue

            # EResult 84: rate limit — ждём и прерываем текущий ордер
            if res.get("steam_result") in RATELIMIT_ERESULT_CODES:
                print(f"  [!!] EResult 84: Rate limit. Пауза {DELAY_ON_RATELIMIT}с...")
                time.sleep(DELAY_ON_RATELIMIT)
                break

            if res.get("no_retry"):
                print(f"  [SKIP] {err}")
                break

            print(f"  [ERR] Попытка {attempt}/{retries}: {err}")
            if attempt < retries:
                time.sleep(delay)

        if not placed:
            results.append({**vars(order), "order_id": None, "status": "FAILED"})

        if i < total:
            time.sleep(delay)

    _print_summary(results, currency)
    return results


def _print_summary(results: list, currency: int):
    ok  = [r for r in results if r.get("status") == "OK"]
    bad = [r for r in results if r.get("status") != "OK"]
    print(f"\n{'='*58}")
    print(f"  Итог: {len(ok)}/{len(results)} успешно | {len(bad)} ошибок")
    print(f"{'='*58}")
    if bad:
        print("\n  Не выставлены:")
        for r in bad:
            print(f"    - {r['market_hash_name']} [{r['status']}]")


# =============================================================
#  ENTRY POINT
# =============================================================

def main():
    ap = argparse.ArgumentParser(
        description="Steam Market — авто-выставление buy-ордеров с поддержкой профилей.",
        formatter_class=argparse.RawTextHelpFormatter
    )

    # Профиль
    ap.add_argument("--user-profile", "-u", default="default",
                    help="Имя профиля аккаунта (default). Пример: --user-profile main")
    ap.add_argument("--reauth", action="store_true",
                    help="Принудительная переавторизация через браузер")
    ap.add_argument("--list-profiles", action="store_true",
                    help="Показать все сохранённые профили")

    # Ордера
    ap.add_argument("--excel", default=DEFAULT_EXCEL_FILE,
                    help=f"Excel файл с ордерами (по умолч.: {DEFAULT_EXCEL_FILE})")
    ap.add_argument("--appid", type=int, default=DEFAULT_APPID,
                    help="AppID игры (252490=Rust, 730=CS2)")
    ap.add_argument("--currency", type=int, default=0,
                    help="Переопределить валюту (0 = из профиля)")

    # Запросы
    ap.add_argument("--delay",       type=float, default=DELAY_BETWEEN_ORDERS)
    ap.add_argument("--timeout",     type=int,   default=REQUEST_TIMEOUT)
    ap.add_argument("--retries",     type=int,   default=MAX_RETRIES)
    ap.add_argument("--max-orders",  type=int,   default=MAX_ORDERS_PER_RUN,
                    help="Макс. ордеров за один запуск (0 = все)")

    # Прокси
    ap.add_argument("--proxy-scheme", default="http")
    ap.add_argument("--no-proxy",    action="store_true")

    # Прочее
    ap.add_argument("--skip-validate", action="store_true",
                    help="Пропустить сетевую проверку сессии")
    ap.add_argument("-v", "--verbose", action="store_true")

    args = ap.parse_args()

    # --- Список профилей ---
    if args.list_profiles:
        profiles = list_profiles()
        if not profiles:
            print("Нет сохранённых профилей.")
        else:
            print("Сохранённые профили:")
            for p in profiles:
                data = load_profile(p)
                sls = data.get("steam_login_secure", "")
                valid, secs = check_token_expiry(sls)
                status = f"JWT OK (~{secs//86400}д)" if valid else "JWT ИСТЁК"
                currency = data.get("currency", "?")
                cur_name = CURRENCY_NAMES.get(currency, str(currency))
                print(f"  [{p}]  валюта={cur_name}  {status}")
        return

    # --- Получение сессии ---
    session_id, steam_login_secure, currency, all_cookies, user_agent = ensure_session(
        profile=args.user_profile,
        force_reauth=args.reauth,
        skip_validate=args.skip_validate,
    )

    # Переопределение валюты из CLI
    if args.currency:
        currency = args.currency

    # --- Ордера ---
    orders = load_orders_from_excel(args.excel)
    if not orders:
        print("[!] Нет ордеров. Заполни orders.xlsx и запусти снова.")
        return

    if args.appid != DEFAULT_APPID:
        for o in orders:
            o.appid = args.appid

    # --- Прокси ---
    proxies = []
    if not args.no_proxy and USE_PROXIES:
        proxies = load_proxies(DEFAULT_PROXIES_FILE)

    # --- Запуск ---
    run(
        orders=orders,
        session_id=session_id,
        steam_login_secure=steam_login_secure,
        currency=currency,
        proxies=proxies,
        proxy_scheme=args.proxy_scheme,
        delay=args.delay,
        retries=args.retries,
        timeout=args.timeout,
        max_orders=args.max_orders,
        verbose=args.verbose,
        profile=args.user_profile,
        all_cookies=all_cookies,
        user_agent=user_agent,
    )


if __name__ == "__main__":
    main()
